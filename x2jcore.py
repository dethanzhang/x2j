# !/usr/bin/python3
# -*- coding: utf-8 -*-
# author: dethanz
# contact: dethanzhang@hotmail.com
# 需要安装openpyxl

import os
import x2jutils
from openpyxl import load_workbook


class x2jcore:
    def __init__(self):
        self.title_pos = 2
        self.type_pos = 3
        self.subtype_pos = 4
        self.content_begin = 6
        self.excel_path = "配置表"
        self.output_path = "output_temp"
        self.save_path = "json"
        self.folder_keys = None
        self.folder_dict = None
        self.single_folder = None

    def storeErrorMsg(self, i, j):
        quotient, remainder = divmod(j, 26)
        # 0~25 对应A~Z, 26~51 对应AA~AZ, 52~77 对应BA~BZ
        if quotient == 0:
            col_name = chr(remainder + 65)
        else:
            col_name = chr(quotient + 64) + chr(remainder + 65)
        self.error_msg[self.current_sheet].append(
            "行【%s】 【%s】列 填写错误" % (i + self.content_begin, col_name)
        )
        self.error_cnt += 1

    def outputJson(self, sheet_name, table):
        if self.folder_keys and sheet_name in self.folder_keys:
            folder_path = os.path.join(self.output_path, *self.folder_dict[sheet_name])
            os.makedirs(folder_path, exist_ok=True)

        elif self.single_folder:
            folder_path = os.path.join(self.output_path, *self.single_folder)
            os.makedirs(folder_path, exist_ok=True)
            self.single_folder = None

        else:
            folder_path = self.output_path
        x2jutils.writeJsonFile(
            os.path.join(folder_path, sheet_name + ".json"),
            table,
        )

    def start(self, filename):  # 返回0=导表成功 1=字段名带空格
        self.error_msg = (
            {}
        )  # 格式为{sheetname1: [错误信息1,...],sheetname2: [错误信息1,...]}
        self.error_cnt = 0
        self.wb = load_workbook(filename, data_only=True)

        try:
            self.readFolderTags(self.wb["输出目录"])
            print("已读取目录, 将按照设定目录结构导出")
        except:
            pass

        for sheet in self.wb.sheetnames:
            if sheet == "":
                continue
            outputs = sheet.split("|")
            if len(outputs) <= 1:
                continue
            sheet_name = outputs[0]
            singleOutput = 0
            if sheet_name.startswith("#"):
                singleOutput = 1
                sheet_name = sheet_name[1:]
            if sheet_name.startswith("^"):
                singleOutput = 2
                sheet_name = sheet_name[1:]
            if sheet_name.startswith("$"):
                singleOutput = 3
                sheet_name = sheet_name[1:]
            if sheet_name.startswith("%"):
                singleOutput = 4
                sheet_name = sheet_name[1:]

            marks = outputs[1].split("@")
            if len(marks) > 1:
                self.single_folder = marks[1:]

            self.current_sheet = sheet
            self.sheet_data = self.wb[sheet]
            self.titles = [
                cell.value for cell in self.sheet_data[self.title_pos]
            ]  # 拿到表头
            self.types = [
                cell.value for cell in self.sheet_data[self.type_pos]
            ]  # 拿到类型
            self.subTypes = [
                cell.value for cell in self.sheet_data[self.subtype_pos]
            ]  # 拿到子类型备注
            self.max_col = x2jutils.getValidLength(self.titles)

            self.sheet_data = [
                row
                for row in self.sheet_data.iter_rows(
                    min_row=self.content_begin, max_col=self.max_col, values_only=True
                )
                if any(cell is not None for cell in row)
            ]  # 去掉空行仅保留有效数据

            find = -1
            findGroup = -1

            self.error_msg[sheet] = []
            print("导出子表格 >>>>> ", sheet_name)
            for j in range(0, len(self.titles)):
                if self.titles[j] is None:
                    self.titles[j] = "#"
                    continue
                title = self.titles[j]
                if " " in title:
                    self.error_msg[self.current_sheet].append(
                        "%s表中的%s字段包含有空格, 请检查" % (self.current_sheet, title)
                    )
                    self.error_cnt += 1
                    return 1
                if title.startswith("*"):
                    self.titles[j] = title[1:]
                    find = j
                if title.startswith("^"):
                    self.titles[j] = title[1:]
                    findGroup = j

            # 单独输出sheet为整个json文件 并以id为key包含每行内容为对象进行序列化
            if singleOutput == 1:
                if findGroup != -1:
                    if find != -1:
                        table = self.readExcelByGroup(findGroup, find)
                        self.outputJson(sheet_name, table)
                elif find != -1:
                    table = self.readExcelByKey(find)
                    self.outputJson(sheet_name, table)
                else:
                    # 都没有找到 为无key文件
                    table = self.readExcelNoKey()
                    self.outputJson(sheet_name, table)
                continue

            if singleOutput == 2:
                if findGroup != -1:
                    table = self.readExcelWithGroup(findGroup)
                    self.outputJson(sheet_name, table)
                continue

            if singleOutput == 3:
                try:
                    self.readLocalizationExcel(self.wb["异常字符集"])
                    print("已读取错误字符集")
                except KeyError:
                    self.readLocalizationExcel()
                continue

        return 0

    # 将单个sheet输出为一个数组, 一行数据为一个字典对象
    def readExcelNoKey(self):
        table = []
        for i, row in enumerate(self.sheet_data):
            if row[0] is None or str(row[0]).startswith(
                "#"
            ):  # 首列内容为空或包含#号则表示本行不导出
                continue
            line = {}
            for j in range(0, self.max_col):
                if self.titles[j].startswith("#"):
                    continue
                if self.types[j] != "":
                    line[self.titles[j]] = x2jutils.getValueByType(
                        row[j], self.types[j], self.subTypes[j]
                    )
                    if line[self.titles[j]] == None:
                        self.storeErrorMsg(i, j)
            table.append(line)
        return table

    # 将单个sheet输出为一个数组, 根据标记位置, 将多行组成一个数组. 只支持1个子层级,即[{ 主键 [{从属内容},] },]
    def readExcelWithGroup(self, findGroup):
        table = []
        jumpFlag = False
        for i, row in enumerate(self.sheet_data):
            line = {}

            if row[0]:  # 首列不为空,表示有新增主键
                if str(row[0]).startswith(
                    "#"
                ):  # 首列内容包含#号则表示本id的所有内容不导出
                    jumpFlag = True
                    continue
                jumpFlag = False
                try:  # 避免首行报错
                    table.append(alevel)
                except NameError:
                    pass
                alevel = {}
                for j in range(0, findGroup):  # 创建主键内容
                    if self.titles[j].startswith("#"):
                        continue
                    alevel[self.titles[j]] = x2jutils.getValueByType(
                        row[j], self.types[j], self.subTypes[j]
                    )
                    if alevel[self.titles[j]] is None:  # 处理报错
                        print(
                            x2jutils.getValueByType(
                                row[j], self.types[j], self.subTypes[j], debug=True
                            )
                        )
                        self.storeErrorMsg(i, j)
                alevel[self.titles[findGroup]] = []
            if jumpFlag:
                continue
            for j in range(findGroup + 1, self.max_col):  # 将每行的从属内容合并
                if self.titles[j].startswith("#"):
                    continue
                line[self.titles[j]] = x2jutils.getValueByType(
                    row[j], self.types[j], self.subTypes[j]
                )
                if line[self.titles[j]] is None:  # 处理报错
                    print(
                        x2jutils.getValueByType(
                            row[j], self.types[j], self.subTypes[j], debug=True
                        )
                    )
                    self.storeErrorMsg(i, j)
            alevel[self.titles[findGroup]].append(line)
        table.append(alevel)
        return table

    # # 将单个sheet输出为一个字典, 根据标记位置的列作为字典key. 每一行的其它内容再嵌套成一个字典. {id1:{title1:v1,title2:v2...},...}
    # def readExcelByKey(self, find):
    #     table = {}
    #     main_col = [cell.value for cell in self.sheet_data[find + 1]]
    #     main_type = self.types[find]
    #     for i in range(self.content_begin, len(main_col)):
    #         row = [cell.value for cell in self.sheet_data[i + 1]]
    #         if main_col[i] != "":
    #             id = x2jutils.getValueByType(main_col[i], main_type)
    #             table[id] = {}
    #             for k in range(0, len(self.types)):
    #                 if self.types[k] != "" and row[k] != "":
    #                     value = x2jutils.getValueByType(
    #                         row[k], self.types[k], self.subTypes[k]
    #                     )
    #                     key = self.titles[k]
    #                     table[id][key] = value
    #     return table

    # # 预留类型
    # def readExcelByGroup(self, findGroup, find):
    #     table = {}
    #     main_group_col = [cell.value for cell in self.sheet_data[findGroup + 1]]
    #     main_col = [cell.value for cell in self.sheet_data[find + 1]]
    #     main_group_type = self.types[findGroup]
    #     main_type = self.types[find]
    #     for i in range(self.content_begin, len(main_group_col)):
    #         row = [cell.value for cell in self.sheet_data[i + 1]]
    #         if main_group_col[i] != "":
    #             group = x2jutils.getValueByType(main_group_col[i], main_group_type)
    #             if group not in table:
    #                 table[group] = {}
    #             id = x2jutils.getValueByType(main_col[i], main_type)
    #             table[group][id] = {}
    #             for k in range(0, len(self.types)):
    #                 if self.types[k] != "" and row[k] != "":
    #                     value = x2jutils.getValueByType(
    #                         row[k], self.types[k], self.subTypes[k]
    #                     )
    #                     key = self.titles[k]
    #                     table[group][id][key] = value
    #     return table

    # **多语言文本表** 读取单个sheet, 以首列为字典key, 后续每一列作为不同dict的value, 并以每一列的id作为json名输出
    def readLocalizationExcel(self, char_data=None):
        if char_data:
            list_badChar = [
                cell.value for cell in char_data["A"] if cell.value is not None
            ][1:]
            list_goodChar = [
                cell.value for cell in char_data["C"] if cell.value is not None
            ][1:]

        self.sheet_data = list(zip(*self.sheet_data))
        keys = self.sheet_data[0]
        for j in range(1, self.max_col):
            if self.titles[j].startswith("#"):
                continue
            col = list(self.sheet_data[j])
            if char_data:
                col = x2jutils.fixBadChar(
                    col,
                    list_badChar,
                    list_goodChar,
                )
            table = dict(zip(keys, col[: len(keys)]))
            x2jutils.writeJsonFile(
                os.path.join(self.output_path, self.titles[j] + ".json"), table
            )

    # **读取输出目录页面** 读取单个sheet, 以首行作为路径名称字典key, 后续每一行作为不同dict的value, 并以每一列的id作为json名输出
    def readFolderTags(self, sheet_data):
        sheet_data = [
            row
            for row in sheet_data.iter_rows(min_row=1, values_only=True)
            if any(cell is not None for cell in row)
        ]  # 去掉空行仅保留有效数据

        self.folder_dict = {}
        self.folder_keys = set()
        for tup in sheet_data:
            key = tup[-1]  # 最后一个元素作为键
            self.folder_keys.add(key)
            values = list(tup[:-1])  # 其余元素作为值列表
            self.folder_dict[key] = values


if __name__ == "__main__":
    ax = x2jcore()
    # 获得所有json目录路径
    all_json_path = x2jutils.getAllFolders(ax.save_path)

    # 获得所有excel文件路径
    all_xlsx_path = []
    xlsx_folder_list = x2jutils.getAllFolders(ax.excel_path)
    for a in xlsx_folder_list:
        all_xlsx_path += [os.path.join(a, b) for b in x2jutils.xlsxFileList(a)]

    # 清理temp目录
    x2jutils.clearTempFiles(ax.output_path)

    # 导表
    for i in range(0, len(all_xlsx_path)):
        print(i + 1, "---", all_xlsx_path[i])
    print("请输入要导出的表的序号:")
    selectFile = int(input())
    if selectFile > len(all_xlsx_path) or selectFile < 1:
        print("输入错误 请重新执行脚本")
    else:
        ax.start(all_xlsx_path[selectFile - 1])
